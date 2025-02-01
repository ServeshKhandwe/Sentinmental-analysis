import streamlit as st
import feedparser
import requests
import pandas as pd
import plotly.graph_objs as go
import time
from transformers import pipeline
import yfinance as yf
from streamlit.components.v1 import html as st_html
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import unicodedata

# -----------------------------
# Alpha Vantage & Model Config
# -----------------------------
ALPHA_VANTAGE_API_KEY = st.secrets["APIKEY"] 
ALPHA_VANTAGE_API_URL = 'https://www.alphavantage.co/query'

MODEL_OPTIONS = {
    "FinBERT (ProsusAI/finbert)": "ProsusAI/finbert",
    "RoBERTa-Large English (siebert/sentiment-roberta-large-english)": "siebert/sentiment-roberta-large-english",
    "Twitter RoBERTa (cardiffnlp/twitter-roberta-base-sentiment)": "cardiffnlp/twitter-roberta-base-sentiment",
    "Multilingual BERT (nlptown/bert-base-multilingual-uncased-sentiment)": "nlptown/bert-base-multilingual-uncased-sentiment",
    "DistilBERT Emotion (distilbert-base-uncased-emotion)": "j-hartmann/emotion-english-distilroberta-base"
}

# -----------------------------
# Scrolling Ticker Functions
# -----------------------------
stock_tickers = [
    'AAPL', 'GOOGL', 'MSFT', 'AMZN', 'META', 'NVDA', 'TSLA',
    'JPM', 'JNJ', 'V', 'WMT', 'PG', 'XOM', 'HD', 'BAC',
    'DIS', 'NFLX', 'INTC', 'KO', 'PEP', 'CSCO', 'VZ',
    'MCD', 'ADBE', 'CMCSA'
]

def get_stock_prices(tickers):
    """
    Fetches the most recent closing price for each ticker from Yahoo Finance.
    """
    prices = {}
    for ticker in tickers:
        stock = yf.Ticker(ticker)
        todays_data = stock.history(period='1d')
        if not todays_data.empty:
            prices[ticker] = todays_data['Close'][0]
    return prices

def display_scrolling_ticker(prices):
    """
    Generates and displays an HTML-based ticker using CSS animations.
    """
    ticker_text = ' | '.join([f"{ticker}: ${price:.2f}" for ticker, price in prices.items()])
    html_code = f"""
    <div class="ticker-container">
      <div class="ticker-text">
        {ticker_text}
      </div>
    </div>
    """
    st.markdown(html_code, unsafe_allow_html=True)

# -----------------------------
# Stock & Sentiment Functions
# -----------------------------
def fetch_stock_data(symbol):
    """
    Fetches intraday stock data from Alpha Vantage.
    """
    params = {
        'function': 'TIME_SERIES_INTRADAY',
        'symbol': symbol,
        'interval': '5min',
        'apikey': ALPHA_VANTAGE_API_KEY
    }
    response = requests.get(ALPHA_VANTAGE_API_URL, params=params)
    data = response.json()
    if 'Time Series (5min)' in data:
        df = pd.DataFrame.from_dict(data['Time Series (5min)'], orient='index', dtype=float)
        df = df.rename(columns={
            '1. open': 'Open',
            '2. high': 'High',
            '3. low': 'Low',
            '4. close': 'Close',
            '5. volume': 'Volume'
        })
        df.index = pd.to_datetime(df.index)
        df.sort_index(inplace=True)
        return df
    else:
        st.error("Error fetching stock data. Please check the ticker symbol and try again.")
        return None

def plot_stock_data(df, symbol):
    """
    Plots the stock's closing price over time using Plotly.
    """
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df.index, y=df['Close'], mode='lines', name='Close Price'))
    fig.update_layout(
        title=f'Real-Time Stock Price for {symbol}',
        xaxis_title='Time',
        yaxis_title='Price (USD)',
        xaxis_rangeslider_visible=False,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font_color='#F5F5F5'
    )
    st.plotly_chart(fig, use_container_width=True)

def get_sp500_tickers():
    """Fetches a static list of S&P 500 tickers."""
    tickers = [
        "META", "UNH", "AAPL", "MSFT", "AMZN", "NVDA", "GOOGL", "TSLA", "GOOG",
        "JNJ", "JPM", "V", "PG", "MA", "HD", "CVX", "LLY", "ABBV", "PEP"
    ]
    return tickers

# -----------------------------
# Utility Functions for Text & Excel
# -----------------------------
def clean_text(text):
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")

def shorten_title(title, length=80):
    return title if len(title) <= length else title[:length] + "..."

def create_excel_with_formatting(ticker, articles):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker[:30]} Analysis"  # Truncate title if necessary

    # Title above the columns
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Sentiment Analysis for {ticker[:30]}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Two empty rows after the title
    ws.append([])
    ws.append([])

    # Column headers
    headers = ["Title", "Sentiment", "Published", "Summary", "Link"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data with sentiment-based row coloring
    sorted_articles = sorted(articles, key=lambda x: (x["sentiment"], x["published"]), reverse=True)
    for article in sorted_articles:
        sentiment = article["sentiment"]
        color = (
            "d4edda" if sentiment == "positive" else "f8d7da" if sentiment == "negative" else "fff3cd"
        )
        ws.append([
            article["title"],
            sentiment.capitalize(),
            article["published"],
            article["summary"],
            article["link"][:50] + "..."  # Truncate the link for display
        ])
        row_num = ws.max_row
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=12)

    # Add sentiment count summary
    ws.append([])
    ws.append([])
    ws.append(["Sentiment", "Count"])
    ws.append(["Positive", sum(1 for article in articles if article["sentiment"] == "positive")])
    ws.append(["Neutral", sum(1 for article in articles if article["sentiment"] == "neutral")])
    ws.append(["Negative", sum(1 for article in articles if article["sentiment"] == "negative")])
    for row in range(ws.max_row - 3, ws.max_row + 1):
        for col in range(1, 3):
            ws.cell(row=row, column=col).font = Font(bold=(row == ws.max_row - 3), size=12)

    # Set uniform column widths
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 40.0

    return wb

# -----------------------------
# Main Streamlit App
# -----------------------------
def main():
    # Set wide layout and custom page title
    st.set_page_config(page_title="Integrated Stock Sentiment App", layout="wide")

    # --- Inject Custom CSS for Modern Look & Article Boxes ---
    st.markdown(
        """
        <style>
        /* Import a modern font from Google Fonts */
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
        html, body {
            background: linear-gradient(120deg, #1C1C1C, #2E2E2E);
            font-family: 'Poppins', sans-serif;
            color: #F5F5F5;
        }
        .block-container {
            background-color: transparent !important;
            padding: 2rem !important;
        }
        h1, h2, h3, h4, h5 {
            color: #00FFC8;
            margin-top: 1rem;
        }
        .ticker-container {
            background-image: linear-gradient(to right, #06beb6, #48b1bf);
            color: #fff;
            overflow: hidden;
            white-space: nowrap;
            box-sizing: border-box;
            padding: 10px 0;
            margin-bottom: 20px;
            border-radius: 10px;
        }
        .ticker-text {
            display: inline-block;
            padding-left: 100%;
            font-size: 1.2rem;
            font-weight: bold;
            animation: ticker 30s linear infinite;
        }
        @keyframes ticker {
          0%   { transform: translate3d(0, 0, 0); }
          100% { transform: translate3d(-100%, 0, 0); }
        }
        .stButton>button {
            background-color: #00FFC8;
            color: #1C1C1C;
            border-radius: 8px;
            padding: 0.6em 1.2em;
            font-weight: 600;
            border: none;
            transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out;
        }
        .stButton>button:hover {
            background-color: #48b1bf;
            color: #fff;
            cursor: pointer;
        }
        .content-card {
            background-color: rgba(255, 255, 255, 0.05);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 10px;
            padding: 1.5rem;
            margin-bottom: 2rem;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.3);
        }
        .sentiment-box {
            background-color: #1c1c1c;
            border-radius: 8px;
            padding: 1rem;
            margin-top: 1rem;
            color: #00FFC8;
            font-weight: 600;
        }
        /* Article box styling */
        .article-box {
            padding: 10px;
            margin-bottom: 15px;
            border-radius: 5px;
            border: 2px solid;
        }
        .positive {
            border-color: #28a745;
        }
        .neutral {
            border-color: #ffc107;
        }
        .negative {
            border-color: #dc3545;
        }
        .article-title {
            font-size: 16px;
            font-weight: bold;
        }
        .sentiment-arrow {
            font-size: 14px;
            font-weight: bold;
            display: inline-block;
            margin-right: 5px;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.title("Stock Sentiment Analysis App with Real-Time Ticker & Excel Export")

    # --- Ticker Section ---
    st.write("### Live Stock Prices")
    with st.container():
        prices = get_stock_prices(stock_tickers)
        display_scrolling_ticker(prices)

    st.write("---")

    # --- Sentiment Analysis Controls ---
    st.write("### Sentiment Analysis Controls")
    selected_model = st.selectbox("Choose a sentiment model", list(MODEL_OPTIONS.keys()))
    use_sp500 = st.checkbox("Choose ticker from S&P 500 list", value=True)
    if use_sp500:
        sp500_list = get_sp500_tickers()
        default_index = sp500_list.index("META") if "META" in sp500_list else 0
        ticker = st.selectbox("Select Stock Ticker", sp500_list, index=default_index)
    else:
        ticker = st.text_input("Enter Stock Ticker (e.g., META):", "META")
    keyword = st.text_input("Enter Keyword for Filtering Articles:", ticker.lower())
    
    # Option to export results as Excel
    export_excel = st.checkbox("Export Results to Excel", value=False)
    # Option to sort articles
    sort_option = st.radio("Sort Articles By:", ["Date", "Sentiment"], index=0, horizontal=True)

    # --- Analysis Button ---
    if st.button("Analyze"):
        st.write("## Results")

        # 1. Load selected sentiment analysis model
        model_name = MODEL_OPTIONS[selected_model]
        with st.spinner("Loading sentiment analysis model..."):
            pipe = pipeline("text-classification", model=model_name)

        # 2. Fetch stock data & display chart
        with st.spinner(f"Fetching stock data for {ticker}..."):
            stock_data = fetch_stock_data(ticker)
            if stock_data is not None:
                with st.container():
                    st.write(f"### Intraday Stock Chart for {ticker}")
                    plot_stock_data(stock_data, ticker)

        # 3. Fetch RSS feed & run sentiment analysis on news articles
        rss_url = f"https://finance.yahoo.com/rss/headline?s={ticker}"
        with st.spinner("Fetching and analyzing news articles..."):
            feed = feedparser.parse(rss_url)
            articles_list = []
            total_score = 0
            num_articles = 0
            start_time = time.time()

            for entry in feed.entries:
                # Filter based on keyword in title or summary
                if keyword.lower() not in entry.summary.lower() and keyword.lower() not in entry.title.lower():
                    continue

                sentiment = pipe(entry.summary)[0]
                label = sentiment["label"]
                score = sentiment["score"]

                # Map sentiment label to one of positive/negative/neutral
                if "pos" in label.lower():
                    mapped_label = "positive"
                    total_score += score
                    num_articles += 1
                elif "neg" in label.lower():
                    mapped_label = "negative"
                    total_score -= score
                    num_articles += 1
                else:
                    mapped_label = "neutral"

                articles_list.append({
                    "title": entry.title,
                    "published": entry.published,
                    "summary": entry.summary,
                    "link": entry.link,
                    "sentiment": mapped_label,
                    "score": score
                })

            elapsed_time = time.time() - start_time

        # 4. Overall sentiment summary
        if num_articles > 0:
            if total_score > 0.15:
                overall_sentiment = "Positive"
            elif total_score < -0.15:
                overall_sentiment = "Negative"
            else:
                overall_sentiment = "Neutral"

            st.markdown(f"""
            <div class="sentiment-box">
            <h4>Overall Sentiment: {overall_sentiment}</h4>
            <p>Aggregated Score: {total_score:.2f}</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.error("No articles found matching the keyword.")

        # 5. Sentiment distribution chart using Plotly in a centered column
        positive_count = sum(1 for a in articles_list if a["sentiment"] == "positive")
        neutral_count = sum(1 for a in articles_list if a["sentiment"] == "neutral")
        negative_count = sum(1 for a in articles_list if a["sentiment"] == "negative")
        labels = ["Positive", "Neutral", "Negative"]
        values = [positive_count, neutral_count, negative_count]
        fig = go.Figure(data=[go.Bar(
            x=labels,
            y=values,
            marker_color=["#28a745", "#ffc107", "#dc3545"]
        )])
        fig.update_layout(
            title=f"Sentiment Distribution for {ticker}",
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font_color="#F5F5F5"
        )

        # Use columns to center the chart
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.plotly_chart(fig, use_container_width=True)


        # 6. Sort articles based on selected option
        if sort_option == "Date":
            articles_list.sort(key=lambda x: x["published"], reverse=True)
        elif sort_option == "Sentiment":
            articles_list.sort(key=lambda x: (x["sentiment"], x["published"]), reverse=True)

        # 7. Display articles with colored sentiment boxes
        st.write(f"### Articles ({len(articles_list)})")
        for article in articles_list:
            sentiment_class = "positive" if article["sentiment"] == "positive" else "negative" if article["sentiment"] == "negative" else "neutral"
            st.markdown(
                f"""
                <div class="article-box {sentiment_class}">
                <div class="article-title">{shorten_title(clean_text(article['title']))}</div>
                <strong>Published:</strong> {article.get('published', 'N/A')}<br>
                <strong>Summary:</strong> {clean_text(article['summary'])}<br>
                <span class="sentiment-arrow">&rarr;</span> {article['sentiment'].capitalize()} ({article['score']:.2f})<br>
                <a href="{article['link']}" target="_blank">Read Full Article</a>
                </div>
                """, unsafe_allow_html=True
            )

        # 8. Excel export functionality
        if export_excel and articles_list:
            wb = create_excel_with_formatting(ticker, articles_list)
            excel_data = BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)
            st.download_button(
                label="Download Analysis (Excel)",
                data=excel_data,
                file_name=f"{ticker}_sentiment_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # 9. Display performance stats
        st.write("---")
        st.write(f"**Model Selected:** {model_name}")
        st.write(f"**Inference Time (seconds):** {elapsed_time:.2f}")

if __name__ == "__main__":
    main()
